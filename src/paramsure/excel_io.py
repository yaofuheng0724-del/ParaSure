from __future__ import annotations

from collections.abc import Iterable
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from .models import ComplianceResult, EvidenceSource, ProductParameter, TenderRequirement, Verdict
from .text import clean_marker, normalize_text


PARAMETER_HEADERS = {
    "描述",
    "详细描述",
    "功能描述",
    "功能参数说明",
    "指标要求",
    "技术要求",
    "技术指标",
    "详细参数",
    "功能指标",
    "具体功能项",
    "指标项",
    "功能项",
    "技术大类",
}
MODULE_HEADERS = {"模块", "类别", "分类", "品类", "能力项", "功能模块", "一级菜单", "指标项", "技术大类"}
FEATURE_HEADERS = {"功能项", "指标项", "子项", "技术指标", "功能指标", "具体功能项", "三级菜单"}
VERSION_HEADERS = {"版本", "版本（AISOC版/企业版/Mini版/Tiny版）"}
REMARK_HEADERS = {"备注", "备注说明", "注意事项", "说明（此列给客户请删除）", "优势"}


def infer_product_name(path: Path) -> str:
    name = path.stem.replace("招投标参数", "").replace("招标参数", "")
    name = name.strip(" -_")
    return name or path.stem


def iter_excel_files(path: Path) -> Iterable[Path]:
    if path.is_file() and path.suffix.lower() == ".xlsx" and not path.name.startswith("~$"):
        yield path
        return
    for file in sorted(path.glob("*.xlsx")):
        if not file.name.startswith("~$"):
            yield file


def merged_cell_value(ws: Worksheet, row: int, col: int) -> Any:
    value = ws.cell(row=row, column=col).value
    if value is not None:
        return value
    for merged_range in ws.merged_cells.ranges:
        if merged_range.min_row <= row <= merged_range.max_row and merged_range.min_col <= col <= merged_range.max_col:
            return ws.cell(merged_range.min_row, merged_range.min_col).value
    return None


def row_values(ws: Worksheet, row: int, max_col: int | None = None) -> list[str]:
    max_col = max_col or ws.max_column
    return [normalize_text(merged_cell_value(ws, row, col)) for col in range(1, max_col + 1)]


def guess_header_row(ws: Worksheet, max_scan_rows: int = 20) -> int:
    best_row = 1
    best_score = -1
    for row in range(1, min(ws.max_row, max_scan_rows) + 1):
        values = row_values(ws, row)
        non_empty = sum(1 for value in values if value)
        header_hits = sum(1 for value in values if value in PARAMETER_HEADERS or value in MODULE_HEADERS or value in FEATURE_HEADERS)
        score = header_hits * 10 + min(non_empty, 8)
        if score > best_score:
            best_row = row
            best_score = score
    return best_row


def _header_map(headers: list[str]) -> dict[str, list[int]]:
    mapping: dict[str, list[int]] = {}
    for idx, header in enumerate(headers):
        if header:
            mapping.setdefault(header, []).append(idx)
    return mapping


def _first_value(row: list[str], headers: list[str], candidates: set[str]) -> str:
    for idx, header in enumerate(headers):
        if header in candidates and idx < len(row) and row[idx]:
            return row[idx]
    return ""


def _collect_description(row: list[str], headers: list[str]) -> str:
    values: list[str] = []
    for idx, header in enumerate(headers):
        if header in PARAMETER_HEADERS and idx < len(row) and row[idx]:
            values.append(row[idx])
    if values:
        return " ".join(dict.fromkeys(values))
    long_values = [value for value in row if len(value) >= 12]
    return max(long_values, key=len, default="")


def _is_noise_row(row: list[str]) -> bool:
    joined = "".join(row)
    if not joined:
        return True
    if joined.startswith("版本：") or joined.startswith("说明："):
        return True
    return False


def load_product_parameters(path: Path) -> list[ProductParameter]:
    wb = load_workbook(path, data_only=True)
    product = infer_product_name(path)
    parameters: list[ProductParameter] = []

    for sheet_name in wb.sheetnames:
        if any(keyword in sheet_name for keyword in ("更新", "日志", "SLA")):
            continue
        ws = wb[sheet_name]
        header_row = guess_header_row(ws)
        headers = row_values(ws, header_row)
        if not any(headers):
            continue
        last_module = ""
        last_feature = ""
        for row_number in range(header_row + 1, ws.max_row + 1):
            row = row_values(ws, row_number, len(headers))
            if _is_noise_row(row):
                continue
            module = _first_value(row, headers, MODULE_HEADERS) or last_module
            feature = _first_value(row, headers, FEATURE_HEADERS) or last_feature
            description = _collect_description(row, headers)
            version = _first_value(row, headers, VERSION_HEADERS)
            remarks = _first_value(row, headers, REMARK_HEADERS)
            edition_values = [
                row[idx]
                for idx, header in enumerate(headers)
                if idx < len(row) and row[idx] and any(mark in header for mark in ("版", "型号"))
            ]
            edition = " ".join(dict.fromkeys(edition_values))

            if module:
                last_module = module
            if feature:
                last_feature = feature
            if not description and not feature:
                continue
            evidence = " ".join(part for part in [module, feature, description, remarks] if part)
            if len(evidence) < 6:
                continue
            parameters.append(
                ProductParameter(
                    product=product,
                    module=clean_marker(module),
                    feature=clean_marker(feature),
                    description=clean_marker(description),
                    version=version,
                    edition=edition,
                    remarks=remarks,
                    source_file=str(path),
                    sheet_name=sheet_name,
                    row_number=row_number,
                    raw={header or f"col_{idx + 1}": value for idx, (header, value) in enumerate(zip(headers, row)) if value},
                )
            )
    return parameters


def load_tender_requirements(path: Path) -> list[TenderRequirement]:
    wb = load_workbook(path, data_only=True)
    requirements: list[TenderRequirement] = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        header_row = guess_header_row(ws)
        headers = row_values(ws, header_row)
        if not any(headers):
            headers = [f"col_{i}" for i in range(1, ws.max_column + 1)]
            header_row = 0
        for row_number in range(header_row + 1, ws.max_row + 1):
            row = row_values(ws, row_number, len(headers))
            if _is_noise_row(row):
                continue
            description = _collect_description(row, headers)
            title = _first_value(row, headers, FEATURE_HEADERS | MODULE_HEADERS)
            if not description:
                long_values = [value for value in row if len(value) >= 8]
                description = max(long_values, key=len, default="")
            if not description and not title:
                continue
            requirement_id = _first_value(row, headers, {"序号", "编号"}) or f"{sheet_name}-{row_number}"
            requirements.append(
                TenderRequirement(
                    requirement_id=requirement_id,
                    title=clean_marker(title),
                    description=clean_marker(description),
                    source_file=str(path),
                    sheet_name=sheet_name,
                    row_number=row_number,
                    raw={header or f"col_{idx + 1}": value for idx, (header, value) in enumerate(zip(headers, row)) if value},
                )
            )
    return requirements


def write_results(path: Path, results: list[ComplianceResult]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "符合性矩阵"
    headers = [
        "客户参数ID",
        "客户参数原文",
        "指定产品",
        "符合性结论",
        "置信度",
        "匹配功能项",
        "证据来源",
        "证据文本摘要",
        "证据位置",
        "Web截图路径",
        "API响应摘要",
        "风险备注",
        "建议应答口径",
    ]
    ws.append(headers)
    for result in results:
        ws.append(
            [
                result.requirement.requirement_id,
                result.requirement.text,
                result.product,
                result.verdict.value if isinstance(result.verdict, Verdict) else result.verdict,
                round(result.confidence, 3),
                result.matched_feature,
                result.evidence_source.value if isinstance(result.evidence_source, EvidenceSource) else result.evidence_source,
                result.evidence_summary,
                result.evidence_location,
                result.web_artifact,
                result.api_summary,
                result.risk_note,
                result.response_suggestion,
            ]
        )
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max(max_len + 2, 12), 60)
    ws.auto_filter.ref = ws.dimensions
    wb.save(path)
