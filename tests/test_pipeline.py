from pathlib import Path

from openpyxl import Workbook
from openpyxl import load_workbook

from paramsure.excel_io import load_product_parameters
from paramsure.models import VerificationConfig
from paramsure.pipeline import ParaSurePipeline
from paramsure.store import ParameterStore
from paramsure.verifier import VerificationOutcome


def _make_product_excel(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["模块", "功能项", "描述", "响应"])
    ws.append(["检测能力", "SQL注入检测", "支持检测SQL注入漏洞", ""])
    wb.save(path)


def _make_tender_excel(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["序号", "指标项", "详细描述"])
    ws.append(["1", "SQL注入检测", "要求支持检测SQL注入漏洞"])
    wb.save(path)


def test_pipeline_generates_output(tmp_path: Path) -> None:
    product_excel = tmp_path / "慧鉴-智能源代码审计产品招标参数.xlsx"
    tender_excel = tmp_path / "客户参数.xlsx"
    output_excel = tmp_path / "result.xlsx"
    _make_product_excel(product_excel)
    _make_tender_excel(tender_excel)

    store = ParameterStore(tmp_path / "db.sqlite")
    store.add_parameters(load_product_parameters(product_excel))

    results = ParaSurePipeline(store, tmp_path / "artifacts").evaluate_excel(
        tender_file=tender_excel,
        product="慧鉴-智能源代码审计产品",
        output=output_excel,
    )
    assert results
    assert output_excel.exists()


def test_pipeline_skips_web_when_material_match_is_strong(tmp_path: Path, monkeypatch) -> None:
    product_excel = tmp_path / "慧鉴-智能源代码审计产品招标参数.xlsx"
    tender_excel = tmp_path / "客户参数.xlsx"
    output_excel = tmp_path / "result.xlsx"
    _make_product_excel(product_excel)
    _make_tender_excel(tender_excel)
    store = ParameterStore(tmp_path / "db.sqlite")
    store.add_parameters(load_product_parameters(product_excel))

    class FailingWebVerifier:
        def __init__(self, *args, **kwargs):
            pass

        def verify(self, requirement):
            raise AssertionError("strong material matches must not call web verification")

    monkeypatch.setattr("paramsure.pipeline.WebVerifier", FailingWebVerifier)

    results = ParaSurePipeline(store, tmp_path / "artifacts").evaluate_excel(
        tender_file=tender_excel,
        product="慧鉴-智能源代码审计产品",
        output=output_excel,
        verification=VerificationConfig(enabled=True, base_url="http://example.test"),
    )

    assert results[0].web_artifact == ""


def test_pipeline_preserves_web_evidence_when_web_is_unknown(tmp_path: Path, monkeypatch) -> None:
    product_excel = tmp_path / "雷池- Web应用防火墙招标参数.xlsx"
    tender_excel = tmp_path / "客户参数.xlsx"
    output_excel = tmp_path / "result.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.append(["模块", "功能项", "描述"])
    ws.append(["防护", "SQL注入", "支持SQL注入防护"])
    wb.save(product_excel)

    wb = Workbook()
    ws = wb.active
    ws.append(["序号", "指标项", "详细描述"])
    ws.append(["1", "SSO登录", "要求支持SSO/OIDC登录"])
    wb.save(tender_excel)

    store = ParameterStore(tmp_path / "db.sqlite")
    store.add_parameters(load_product_parameters(product_excel))

    class UnknownWebVerifier:
        def __init__(self, *args, **kwargs):
            pass

        def verify(self, requirement):
            return VerificationOutcome(
                confirmed=False,
                confidence=0.0,
                summary="仅命中关键词，缺少可审计页面上下文",
                artifact="/tmp/screenshot.png",
                evidence_path="/tmp/evidence.json",
            )

    monkeypatch.setattr("paramsure.pipeline.WebVerifier", UnknownWebVerifier)

    results = ParaSurePipeline(store, tmp_path / "artifacts").evaluate_excel(
        tender_file=tender_excel,
        product="雷池- Web应用防火墙",
        output=output_excel,
        verification=VerificationConfig(enabled=True, base_url="http://example.test"),
    )

    assert results[0].verdict.value == "未确认"
    assert results[0].web_artifact == "/tmp/screenshot.png"
    assert results[0].web_evidence == "/tmp/evidence.json"

    sheet = load_workbook(output_excel).active
    headers = [cell.value for cell in sheet[1]]
    row = [cell.value for cell in sheet[2]]
    assert "Web证据包路径" in headers
    assert row[headers.index("Web证据包路径")] == "/tmp/evidence.json"
